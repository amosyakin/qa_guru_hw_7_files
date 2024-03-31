import io
from zipfile import ZipFile
from tests.conftest import archive
from pypdf import PdfReader
from openpyxl import load_workbook
import csv


def test_pdf():
    with ZipFile(archive, 'r') as zip_file:
        with zip_file.open('file_pdf.pdf', 'r') as pdf_file:
            reader = PdfReader(pdf_file)
            assert len(reader.pages) == 1
            assert 'This is a test PDF document.' in reader.pages[0].extract_text()


def test_xlsx():
    with ZipFile(archive, 'r') as zip_file:
        with zip_file.open('file_xlsx.xlsx', 'r') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            print(sheet.max_row)
            assert sheet.max_row == 10
            assert sheet.cell(row=8, column=5).value == 'Great Britain'


def test_csv():
    with (ZipFile(archive, 'r') as zip_file):
        with zip_file.open('file_csv.csv', 'r') as csv_file:
            csv_text = io.TextIOWrapper(csv_file, encoding='utf-8')
            csvreader = csv.reader(csv_text)
            for row in csvreader:
                assert len(row) == 10
